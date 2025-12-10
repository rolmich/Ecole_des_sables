from rest_framework import generics, status, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, F
from django.db import transaction
from django.utils import timezone

from .models import Stage, Participant, Village, Bungalow, Language, ActivityLog, ParticipantStage
from .serializers import (
    StageSerializer, StageCreateSerializer, StageUpdateSerializer, StageListSerializer,
    ParticipantSerializer, ParticipantCreateSerializer, ParticipantUpdateSerializer, ParticipantListSerializer,
    VillageSerializer, VillageListSerializer, BungalowSerializer, BungalowListSerializer,
    LanguageSerializer, LanguageCreateSerializer, LanguageUpdateSerializer, LanguageListSerializer,
    ActivityLogSerializer,
    ParticipantStageSerializer, ParticipantStageCreateSerializer, ParticipantStageUpdateSerializer,
    ParticipantSimpleSerializer, ParticipantCreateSimpleSerializer
)
from .assignment_logic import (
    assign_participant_to_bungalow,
    validate_assignment,
    get_bungalow_availability,
    AssignmentError,
    assign_participants_automatically_for_stage
)
from .activity_logger import (
    log_stage_create, log_stage_update, log_stage_delete,
    log_participant_create, log_participant_update, log_participant_delete,
    log_assignment, log_unassignment,
    log_language_create, log_language_update, log_language_delete,
    log_participant_stage_create, log_participant_stage_delete,
    log_excel_import_participant, log_excel_import_summary,
    log_auto_assignment_individual, log_auto_assignment_summary
)


# ==================== STAGE VIEWS ====================

class StageListCreateView(generics.ListCreateAPIView):
    """Vue pour lister et créer des stages."""

    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'instructor']
    filterset_fields = []
    ordering_fields = ['name', 'start_date', 'end_date', 'created_at']
    ordering = ['-created_at']

    def get_queryset(self):
        """Retourne la queryset des stages."""
        return Stage.objects.all()

    def get_serializer_class(self):
        """Retourne le serializer approprié selon la méthode HTTP."""
        if self.request.method == 'POST':
            return StageCreateSerializer
        return StageListSerializer

    def create(self, request, *args, **kwargs):
        """Créer un stage avec gestion du warning de capacité."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Vérifier si un warning de capacité existe
        capacity_warning = serializer.validated_data.pop('_capacity_warning', None)

        # Sauvegarder le stage
        stage = serializer.save(created_by=self.request.user)

        # Log de l'activité
        log_stage_create(self.request.user, stage)

        # Préparer la réponse
        response_serializer = StageSerializer(stage)
        response_data = response_serializer.data

        # Ajouter le warning si présent
        if capacity_warning:
            overlapping_count = capacity_warning['overlapping_events']
            rooms_needed = capacity_warning['rooms_needed']
            total_rooms_used = capacity_warning['total_rooms_used']
            total_available = capacity_warning['total_available']

            response_data['warning'] = {
                'message': (
                    f"⚠️ ATTENTION: La capacité des chambres est dépassée!\n\n"
                    f"📊 Détails:\n"
                    f"• Chambres nécessaires pour cet événement: {rooms_needed}\n"
                    f"• Nombre d'événements qui se chevauchent: {overlapping_count}\n"
                    f"• Total de chambres utilisées sur cette période: {total_rooms_used}\n"
                    f"• Chambres disponibles au total: {total_available}\n"
                    f"• Dépassement: {total_rooms_used - total_available} chambres\n\n"
                    f"💡 Vous pouvez créer cet événement, mais il faudra ajuster les assignations "
                    f"ou réduire le nombre de participants pour respecter la capacité des chambres."
                ),
                'rooms_needed': rooms_needed,
                'total_rooms_used': total_rooms_used,
                'total_available': total_available,
                'overlapping_events': overlapping_count,
                'deficit': total_rooms_used - total_available
            }

        headers = self.get_success_headers(response_data)
        return Response(response_data, status=status.HTTP_201_CREATED, headers=headers)


class StageRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    """Vue pour récupérer, modifier et supprimer un stage."""

    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Retourne la queryset des stages."""
        return Stage.objects.all()

    def get_serializer_class(self):
        """Retourne le serializer approprié selon la méthode HTTP."""
        if self.request.method in ['PUT', 'PATCH']:
            return StageUpdateSerializer
        return StageSerializer

    def perform_update(self, serializer):
        """Enregistre les modifications et log l'activité."""
        # Sauvegarder les anciennes valeurs
        instance = self.get_object()
        old_data = {
            'name': instance.name,
            'start_date': str(instance.start_date),
            'end_date': str(instance.end_date),
            'event_type': instance.event_type,
            'instructor': instance.instructor,
            'instructor2': instance.instructor2,
            'instructor3': instance.instructor3,
            'capacity': instance.capacity,
            'musicians_count': instance.musicians_count
        }

        # Sauvegarder
        stage = serializer.save()

        # Nouvelles valeurs
        new_data = {
            'name': stage.name,
            'start_date': str(stage.start_date),
            'end_date': str(stage.end_date),
            'event_type': stage.event_type,
            'instructor': stage.instructor,
            'instructor2': stage.instructor2,
            'instructor3': stage.instructor3,
            'capacity': stage.capacity,
            'musicians_count': stage.musicians_count
        }

        # Log de l'activité
        log_stage_update(self.request.user, stage, old_data, new_data)

    def perform_destroy(self, instance):
        """Supprime le stage et log l'activité."""
        stage_name = instance.name
        stage_id = instance.id
        instance.delete()
        # Log de l'activité
        log_stage_delete(self.request.user, stage_name, stage_id)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def stage_statistics(request):
    """Retourne les statistiques des stages."""
    total_stages = Stage.objects.count()
    active_stages = Stage.objects.filter(
        start_date__lte=timezone.now().date(),
        end_date__gte=timezone.now().date()
    ).count()
    upcoming_stages = Stage.objects.filter(
        start_date__gt=timezone.now().date()
    ).count()
    completed_stages = Stage.objects.filter(
        end_date__lt=timezone.now().date()
    ).count()
    
    return Response({
        'total': total_stages,
        'active': active_stages,
        'upcoming': upcoming_stages,
        'completed': completed_stages
    })


# ==================== PARTICIPANT VIEWS ====================

class ParticipantListCreateView(generics.ListCreateAPIView):
    """Vue pour lister et créer des participants."""

    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name', 'last_name', 'email']
    filterset_fields = ['status', 'gender', 'language']
    ordering_fields = ['last_name', 'first_name', 'age', 'created_at']
    ordering = ['last_name', 'first_name']

    def get_queryset(self):
        """Retourne la queryset des participants avec filtres personnalisés."""
        queryset = Participant.objects.prefetch_related('stage_participations', 'languages').select_related('assigned_bungalow').all()

        # Filtre par stage (via ParticipantStage)
        stage_id = self.request.query_params.get('stageId')
        if stage_id:
            queryset = queryset.filter(stage_participations__stage_id=stage_id)

        # Filtre par assignation
        assigned = self.request.query_params.get('assigned')
        if assigned is not None:
            if assigned.lower() == 'true':
                queryset = queryset.filter(assigned_bungalow__isnull=False)
            elif assigned.lower() == 'false':
                queryset = queryset.filter(assigned_bungalow__isnull=True)

        return queryset

    def get_serializer_class(self):
        """Retourne le serializer approprié selon la méthode HTTP."""
        if self.request.method == 'POST':
            return ParticipantCreateSerializer
        return ParticipantListSerializer

    def perform_create(self, serializer):
        """Associe l'utilisateur connecté comme créateur."""
        participant = serializer.save(created_by=self.request.user)
        # Log de l'activité
        log_participant_create(self.request.user, participant)


class ParticipantRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    """Vue pour récupérer, modifier et supprimer un participant."""

    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Retourne la queryset des participants."""
        return Participant.objects.prefetch_related('stage_participations', 'languages').select_related('assigned_bungalow').all()

    def get_serializer_class(self):
        """Retourne le serializer approprié selon la méthode HTTP."""
        if self.request.method in ['PUT', 'PATCH']:
            return ParticipantUpdateSerializer
        return ParticipantSerializer

    def perform_update(self, serializer):
        """Enregistre les modifications et log l'activité."""
        # Sauvegarder les anciennes valeurs
        instance = self.get_object()
        old_data = {
            'first_name': instance.first_name,
            'last_name': instance.last_name,
            'email': instance.email,
            'gender': instance.gender,
            'age': instance.age,
            'status': instance.status
        }

        # Sauvegarder
        participant = serializer.save()

        # Nouvelles valeurs
        new_data = {
            'first_name': participant.first_name,
            'last_name': participant.last_name,
            'email': participant.email,
            'gender': participant.gender,
            'age': participant.age,
            'status': participant.status
        }

        # Log de l'activité
        log_participant_update(self.request.user, participant, old_data, new_data)

    def perform_destroy(self, instance):
        """Supprime le participant et nettoie toutes les données associées."""
        participant_name = instance.full_name
        participant_id = instance.id

        # Nettoyer les lits des bungalows pour toutes les inscriptions du participant
        registrations = ParticipantStage.objects.filter(participant=instance).select_related('assigned_bungalow')

        for registration in registrations:
            if registration.assigned_bungalow and registration.assigned_bed:
                bungalow = registration.assigned_bungalow
                # Libérer le lit
                for bed in bungalow.beds:
                    if bed.get('id') == registration.assigned_bed:
                        occupant = bed.get('occupiedBy')
                        if occupant and (
                            (isinstance(occupant, dict) and occupant.get('registrationId') == registration.id) or
                            occupant == participant_id
                        ):
                            bed['occupiedBy'] = None
                        break
                bungalow.save()
                bungalow.update_occupancy()

        # Supprimer le participant (les inscriptions seront supprimées en cascade)
        instance.delete()

        # Log de l'activité
        log_participant_delete(self.request.user, participant_name, participant_id)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def participant_statistics(request):
    """Retourne les statistiques des participants."""
    total_participants = Participant.objects.count()
    assigned_participants = Participant.objects.filter(
        assigned_bungalow__isnull=False
    ).count()
    unassigned_participants = total_participants - assigned_participants
    
    # Statistiques par statut
    status_stats = {}
    for status, _ in Participant.STATUS_CHOICES:
        count = Participant.objects.filter(status=status).count()
        status_stats[status] = count
    
    # Statistiques par stage
    stage_stats = []
    for stage in Stage.objects.all():
        count = Participant.objects.filter(stages=stage).count()
        stage_stats.append({
            'stage_id': stage.id,
            'stage_name': stage.name,
            'participant_count': count,
            'capacity': stage.capacity,
            'occupancy_rate': round((count / stage.capacity * 100) if stage.capacity > 0 else 0, 2)
        })
    
    return Response({
        'total': total_participants,
        'assigned': assigned_participants,
        'unassigned': unassigned_participants,
        'by_status': status_stats,
        'by_stage': stage_stats
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def assign_participant(request, participant_id):
    """Assigne un participant à un bungalow avec validation complète."""
    try:
        participant = Participant.objects.get(pk=participant_id)
    except Participant.DoesNotExist:
        return Response(
            {
                'error': f'ERREUR: Participant non trouvé (ID: {participant_id}). Vérifiez que le participant existe dans la base de données.',
                'participant_id': participant_id,
                'action': 'assign_participant'
            }, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    bungalow_id = request.data.get('bungalowId')
    bed_id = request.data.get('bed')
    stage_id = request.data.get('stageId')  # ID du stage pour l'assignation
    
    if not bungalow_id:
        return Response(
            {
                'error': 'ERREUR: ID du bungalow manquant. Vous devez fournir un "bungalowId" pour assigner le participant.',
                'participant': f'{participant.first_name} {participant.last_name}',
                'required_fields': ['bungalowId', 'bed', 'stageId']
            }, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if not bed_id:
        return Response(
            {
                'error': 'ERREUR: ID du lit manquant. Vous devez fournir un "bed" (ex: "bed1") pour assigner le participant.',
                'participant': f'{participant.first_name} {participant.last_name}',
                'required_fields': ['bungalowId', 'bed', 'stageId']
            }, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if not stage_id:
        return Response(
            {
                'error': 'ERREUR: ID du stage manquant. Vous devez fournir un "stageId" pour indiquer pour quel stage vous assignez ce participant.',
                'participant': f'{participant.first_name} {participant.last_name}',
                'required_fields': ['bungalowId', 'bed', 'stageId']
            }, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        bungalow = Bungalow.objects.get(pk=bungalow_id)
    except Bungalow.DoesNotExist:
        return Response(
            {
                'error': f'ERREUR: Bungalow non trouvé (ID: {bungalow_id}). Le bungalow demandé n\'existe pas dans la base de données.',
                'bungalow_id': bungalow_id,
                'action': 'assign_participant',
                'suggestion': 'Vérifiez que le bungalow existe ou synchronisez avec "python populate_villages.py"'
            }, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    try:
        stage = Stage.objects.get(pk=stage_id)
    except Stage.DoesNotExist:
        return Response(
            {
                'error': f'ERREUR: Stage non trouvé (ID: {stage_id}). Le stage demandé n\'existe pas.',
                'stage_id': stage_id,
                'action': 'assign_participant'
            }, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Utiliser la logique d'assignation avec toutes les validations
    success, message, details = assign_participant_to_bungalow(
        participant, bungalow, bed_id, stage
    )

    if success:
        # Log de l'activité
        log_assignment(request.user, participant, bungalow, bed_id)

        serializer = ParticipantSerializer(participant)
        return Response({
            'success': True,
            'message': message,
            'participant': serializer.data,
            'details': details
        })
    else:
        return Response(
            {
                'error': message,
                'details': details
            },
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def unassign_participant(request, participant_id):
    """Désassigne un participant de son bungalow."""
    try:
        participant = Participant.objects.get(pk=participant_id)
    except Participant.DoesNotExist:
        return Response(
            {
                'error': f'ERREUR: Participant non trouvé (ID: {participant_id}). Impossible de désassigner un participant inexistant.',
                'participant_id': participant_id,
                'action': 'unassign_participant'
            }, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    if not participant.is_assigned:
        return Response(
            {
                'error': f'ERREUR: Le participant {participant.full_name} n\'est PAS assigné à un bungalow. Impossible de désassigner quelqu\'un qui n\'est pas assigné.',
                'participant': participant.full_name,
                'is_assigned': False
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    old_bungalow = participant.assigned_bungalow
    old_bed = participant.assigned_bed

    participant.assigned_bungalow = None
    participant.assigned_bed = None
    participant.save()

    # Mettre à jour l'occupation du bungalow
    if old_bungalow:
        old_bungalow.update_occupancy()

    # Log de l'activité
    log_unassignment(request.user, participant, old_bungalow, old_bed)

    serializer = ParticipantSerializer(participant)
    return Response({
        'success': True,
        'message': f'Participant {participant.full_name} désassigné du bungalow {old_bungalow.name if old_bungalow else "inconnu"} (lit: {old_bed})',
        'participant': serializer.data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def participants_by_stage(request, stage_id):
    """Retourne tous les participants d'un stage."""
    try:
        stage = Stage.objects.get(pk=stage_id)
    except Stage.DoesNotExist:
        return Response(
            {
                'error': f'ERREUR: Stage non trouvé (ID: {stage_id}). Le stage demandé n\'existe pas dans la base de données.',
                'stage_id': stage_id,
                'action': 'get_participants_by_stage',
                'suggestion': 'Vérifiez l\'ID du stage ou créez-le d\'abord.'
            }, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    participants = Participant.objects.filter(stages=stage)
    serializer = ParticipantListSerializer(participants, many=True)
    
    return Response({
        'success': True,
        'stage': StageSerializer(stage).data,
        'participants': serializer.data,
        'count': participants.count(),
        'message': f'{participants.count()} participant(s) trouvé(s) pour le stage "{stage.name}"'
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def unassigned_participants(request):
    """Retourne les participants non assignés à un bungalow."""
    unassigned = Participant.objects.filter(assigned_bungalow__isnull=True)
    serializer = ParticipantListSerializer(unassigned, many=True)

    return Response({
        'participants': serializer.data,
        'count': unassigned.count()
    })


# ==================== VILLAGE VIEWS ====================

class VillageListView(generics.ListAPIView):
    """Vue pour lister les villages (lecture seule)."""
    
    permission_classes = [IsAuthenticated]
    serializer_class = VillageListSerializer
    
    def get_queryset(self):
        """Retourne tous les villages."""
        return Village.objects.all()


class VillageRetrieveView(generics.RetrieveAPIView):
    """Vue pour récupérer un village avec ses bungalows (lecture seule)."""
    
    permission_classes = [IsAuthenticated]
    serializer_class = VillageSerializer
    
    def get_queryset(self):
        """Retourne tous les villages."""
        return Village.objects.prefetch_related('bungalows').all()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def village_statistics(request):
    """Retourne les statistiques des villages."""
    total_villages = Village.objects.count()
    total_bungalows = Bungalow.objects.count()
    occupied_bungalows = Bungalow.objects.filter(occupancy__gt=0).count()
    empty_bungalows = total_bungalows - occupied_bungalows
    
    # Statistiques par village
    village_stats = []
    for village in Village.objects.all():
        total = village.bungalows.count()
        occupied = village.bungalows.filter(occupancy__gt=0).count()
        village_stats.append({
            'village_name': village.name,
            'total_bungalows': total,
            'occupied_bungalows': occupied,
            'empty_bungalows': total - occupied,
            'occupancy_rate': round((occupied / total * 100) if total > 0 else 0, 2),
            'amenities_type': village.get_amenities_type_display()
        })
    
    return Response({
        'total_villages': total_villages,
        'total_bungalows': total_bungalows,
        'occupied_bungalows': occupied_bungalows,
        'empty_bungalows': empty_bungalows,
        'by_village': village_stats
    })


# ==================== BUNGALOW VIEWS ====================

class BungalowListView(generics.ListAPIView):
    """Vue pour lister les bungalows (lecture seule)."""
    
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name']
    filterset_fields = ['village__name', 'type']
    ordering_fields = ['name', 'village__name', 'occupancy']
    ordering = ['village__name', 'name']
    
    def get_queryset(self):
        """Retourne tous les bungalows avec filtres personnalisés."""
        queryset = Bungalow.objects.select_related('village').all()
        
        # Filtre par village
        village = self.request.query_params.get('village')
        if village:
            queryset = queryset.filter(village__name=village)
        
        # Filtre par disponibilité
        available = self.request.query_params.get('available')
        if available is not None:
            if available.lower() == 'true':
                queryset = queryset.filter(occupancy__lt=F('capacity'))
            elif available.lower() == 'false':
                queryset = queryset.filter(occupancy__gte=F('capacity'))
        
        return queryset
    
    def get_serializer_class(self):
        """Retourne le serializer approprié avec les lits."""
        return BungalowSerializer


class BungalowRetrieveView(generics.RetrieveAPIView):
    """Vue pour récupérer un bungalow (lecture seule)."""
    
    permission_classes = [IsAuthenticated]
    serializer_class = BungalowSerializer
    
    def get_queryset(self):
        """Retourne tous les bungalows."""
        return Bungalow.objects.select_related('village').all()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bungalow_details(request, bungalow_id):
    """Retourne les détails d'un bungalow avec les participants assignés."""
    try:
        bungalow = Bungalow.objects.select_related('village').get(pk=bungalow_id)
    except Bungalow.DoesNotExist:
        return Response(
            {
                'error': f'ERREUR: Bungalow non trouvé (ID: {bungalow_id}). Le bungalow demandé n\'existe pas.',
                'bungalow_id': bungalow_id,
                'action': 'get_bungalow_details',
                'suggestion': 'Vérifiez l\'ID ou synchronisez les bungalows avec "python populate_villages.py"'
            }, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Récupérer les participants assignés à ce bungalow
    participants = Participant.objects.filter(assigned_bungalow=bungalow)

    return Response({
        'success': True,
        'bungalow': BungalowSerializer(bungalow).data,
        'participants': ParticipantListSerializer(participants, many=True).data,
        'message': f'{participants.count()} participant(s) assigné(s) au bungalow {bungalow.name}'
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bungalows_by_village(request, village_name):
    """Retourne tous les bungalows d'un village."""
    try:
        village = Village.objects.get(name=village_name)
    except Village.DoesNotExist:
        available_villages = [v.name for v in Village.objects.all()]
        return Response(
            {
                'error': f'ERREUR: Village "{village_name}" non trouvé. Le village demandé n\'existe pas.',
                'village_name': village_name,
                'available_villages': available_villages,
                'action': 'get_bungalows_by_village',
                'suggestion': f'Villages disponibles: {", ".join(available_villages) if available_villages else "Aucun"}'
            }, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    bungalows = Bungalow.objects.filter(village=village)
    serializer = BungalowSerializer(bungalows, many=True)
    
    return Response({
        'success': True,
        'village': VillageSerializer(village).data,
        'bungalows': serializer.data,
        'count': bungalows.count(),
        'message': f'{bungalows.count()} bungalow(s) trouvé(s) dans le village {village_name}'
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def available_bungalows(request):
    """Retourne tous les bungalows avec des lits disponibles."""
    bungalows = Bungalow.objects.select_related('village').filter(
        occupancy__lt=F('capacity')
    ).order_by('village__name', 'name')

    serializer = BungalowSerializer(bungalows, many=True)

    return Response({
        'bungalows': serializer.data,
        'count': bungalows.count()
    })


# ==================== LANGUAGE VIEWS ====================

class LanguageListCreateView(generics.ListCreateAPIView):
    """Vue pour lister et créer des langues."""

    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'code', 'native_name']
    filterset_fields = ['is_active']
    ordering_fields = ['name', 'display_order', 'created_at']
    ordering = ['display_order', 'name']

    def get_queryset(self):
        """Retourne la queryset des langues."""
        return Language.objects.all()

    def get_serializer_class(self):
        """Retourne le serializer approprié selon la méthode HTTP."""
        if self.request.method == 'POST':
            return LanguageCreateSerializer
        return LanguageListSerializer

    def perform_create(self, serializer):
        """Associe l'utilisateur connecté comme créateur."""
        language = serializer.save(created_by=self.request.user)
        # Log de l'activité
        log_language_create(self.request.user, language)


class LanguageRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    """Vue pour récupérer, modifier et supprimer une langue."""

    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Retourne la queryset des langues."""
        return Language.objects.all()

    def get_serializer_class(self):
        """Retourne le serializer approprié selon la méthode HTTP."""
        if self.request.method in ['PUT', 'PATCH']:
            return LanguageUpdateSerializer
        return LanguageSerializer

    def perform_update(self, serializer):
        """Enregistre les modifications et log l'activité."""
        # Sauvegarder les anciennes valeurs
        instance = self.get_object()
        old_data = {
            'name': instance.name,
            'code': instance.code,
            'native_name': instance.native_name,
            'is_active': instance.is_active,
            'display_order': instance.display_order
        }

        # Sauvegarder
        language = serializer.save()

        # Nouvelles valeurs
        new_data = {
            'name': language.name,
            'code': language.code,
            'native_name': language.native_name,
            'is_active': language.is_active,
            'display_order': language.display_order
        }

        # Log de l'activité
        log_language_update(self.request.user, language, old_data, new_data)

    def perform_destroy(self, instance):
        """Vérifie avant de supprimer une langue."""
        participant_count = instance.participant_count
        if participant_count > 0:
            return Response(
                {
                    'error': f'🚫 IMPOSSIBLE DE SUPPRIMER: Cette langue est utilisée par {participant_count} participant(s). '
                             f'Vous devez d\'abord modifier la langue de ces participants avant de pouvoir la supprimer.',
                    'language_id': instance.id,
                    'language_name': instance.name,
                    'participant_count': participant_count
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        # Sauvegarder les infos avant suppression
        language_name = instance.name
        language_id = instance.id
        instance.delete()
        # Log de l'activité
        log_language_delete(self.request.user, language_name, language_id)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def language_statistics(request):
    """Retourne les statistiques des langues."""
    total_languages = Language.objects.count()
    active_languages = Language.objects.filter(is_active=True).count()
    inactive_languages = Language.objects.filter(is_active=False).count()

    # Langues les plus utilisées
    from django.db.models import Count
    top_languages = Language.objects.annotate(
        usage_count=Count('participant', filter=Q(participant__isnull=False))
    ).order_by('-usage_count')[:5]

    return Response({
        'total': total_languages,
        'active': active_languages,
        'inactive': inactive_languages,
        'topLanguages': [
            {
                'id': lang.id,
                'name': lang.name,
                'code': lang.code,
                'count': lang.usage_count
            }
            for lang in top_languages
        ]
    })


# ==================== ACTIVITY LOG VIEWS ====================

class ActivityLogListView(generics.ListAPIView):
    """Vue pour lister l'historique des activités."""

    permission_classes = [IsAuthenticated]
    serializer_class = ActivityLogSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['description', 'object_repr', 'user__username', 'user__first_name', 'user__last_name']
    filterset_fields = ['user', 'action_type', 'model_name']
    ordering_fields = ['timestamp']
    ordering = ['-timestamp']  # Plus récent en premier
    pagination_class = None  # Désactiver la pagination

    def get_queryset(self):
        """Retourne la queryset des logs d'activité."""
        queryset = ActivityLog.objects.select_related('user').all()

        # Filtre optionnel par utilisateur via query param
        user_id = self.request.query_params.get('user_id', None)
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        # Filtre optionnel par type d'action
        action_type = self.request.query_params.get('action_type', None)
        if action_type:
            queryset = queryset.filter(action_type=action_type)

        # Filtre optionnel par modèle
        model_name = self.request.query_params.get('model_name', None)
        if model_name:
            queryset = queryset.filter(model_name=model_name)

        return queryset


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def activity_log_stats(request):
    """Retourne les statistiques de l'historique d'activité."""
    from django.db.models import Count

    # Total des activités
    total_activities = ActivityLog.objects.count()

    # Activités par type
    activities_by_type = ActivityLog.objects.values('action_type').annotate(
        count=Count('id')
    ).order_by('-count')

    # Activités par modèle
    activities_by_model = ActivityLog.objects.values('model_name').annotate(
        count=Count('id')
    ).order_by('-count')

    # Utilisateurs les plus actifs
    top_users = ActivityLog.objects.values(
        'user__id', 'user__username', 'user__first_name', 'user__last_name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:5]

    # Activités récentes (dernières 24h)
    from datetime import timedelta
    yesterday = timezone.now() - timedelta(days=1)
    recent_activities = ActivityLog.objects.filter(timestamp__gte=yesterday).count()

    return Response({
        'total': total_activities,
        'recent24h': recent_activities,
        'byType': [
            {
                'type': item['action_type'],
                'count': item['count']
            }
            for item in activities_by_type
        ],
        'byModel': [
            {
                'model': item['model_name'],
                'count': item['count']
            }
            for item in activities_by_model
        ],
        'topUsers': [
            {
                'id': user['user__id'],
                'username': user['user__username'],
                'name': f"{user['user__first_name']} {user['user__last_name']}".strip() or user['user__username'],
                'count': user['count']
            }
            for user in top_users
        ]
    })


# ==================== PARTICIPANT STAGE VIEWS ====================

class ParticipantStageListCreateView(generics.ListCreateAPIView):
    """Vue pour lister et créer des inscriptions participant-stage."""

    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['participant__first_name', 'participant__last_name', 'participant__email', 'notes']
    filterset_fields = ['stage', 'participant', 'role']
    ordering_fields = ['created_at', 'arrival_date', 'departure_date']
    ordering = ['-created_at']

    def get_queryset(self):
        """Retourne la queryset des inscriptions."""
        queryset = ParticipantStage.objects.select_related(
            'participant', 'stage', 'participant__assigned_bungalow'
        ).all()

        # Filtre par stage_id dans l'URL
        stage_id = self.kwargs.get('stage_id')
        if stage_id:
            queryset = queryset.filter(stage_id=stage_id)

        return queryset

    def get_serializer_class(self):
        """Retourne le serializer approprié selon la méthode HTTP."""
        if self.request.method == 'POST':
            return ParticipantStageCreateSerializer
        return ParticipantStageSerializer

    def perform_create(self, serializer):
        """Ajoute l'utilisateur créateur et log l'activité."""
        participant_stage = serializer.save(created_by=self.request.user)

        # Log de l'activité
        log_participant_stage_create(
            self.request.user,
            participant_stage.participant,
            participant_stage.stage
        )


class ParticipantStageDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Vue pour récupérer, modifier ou supprimer une inscription."""

    permission_classes = [IsAuthenticated]
    queryset = ParticipantStage.objects.select_related(
        'participant', 'stage', 'participant__assigned_bungalow'
    ).all()

    def get_serializer_class(self):
        """Retourne le serializer approprié selon la méthode HTTP."""
        if self.request.method in ['PUT', 'PATCH']:
            return ParticipantStageUpdateSerializer
        return ParticipantStageSerializer

    def destroy(self, request, *args, **kwargs):
        """Supprime l'inscription, libère le lit et met à jour le compteur du stage."""
        instance = self.get_object()
        stage = instance.stage
        participant_name = instance.participant.full_name
        stage_name = stage.name

        # Libérer le lit du bungalow si l'inscription était assignée
        if instance.assigned_bungalow and instance.assigned_bed:
            bungalow = instance.assigned_bungalow
            for bed in bungalow.beds:
                if bed.get('id') == instance.assigned_bed:
                    occupant = bed.get('occupiedBy')
                    if occupant and (
                        (isinstance(occupant, dict) and occupant.get('registrationId') == instance.id) or
                        occupant == instance.participant_id
                    ):
                        bed['occupiedBy'] = None
                    break
            bungalow.save()
            bungalow.update_occupancy()

        # Supprimer l'inscription
        self.perform_destroy(instance)

        # Mettre à jour le compteur de participants du stage (only role='participant')
        stage.current_participants = stage.participant_registrations.filter(role='participant').count()
        stage.save(update_fields=['current_participants'])

        # Log de l'activité
        log_participant_stage_delete(request.user, participant_name, stage_name)

        return Response(status=status.HTTP_204_NO_CONTENT)


class StageParticipantsView(generics.ListAPIView):
    """Vue pour lister les participants d'un événement spécifique."""

    permission_classes = [IsAuthenticated]
    serializer_class = ParticipantStageSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['participant__first_name', 'participant__last_name', 'participant__email']
    ordering_fields = ['created_at', 'arrival_date', 'participant__last_name']
    ordering = ['participant__last_name', 'participant__first_name']

    def get_queryset(self):
        """Retourne les participants d'un événement."""
        stage_id = self.kwargs.get('stage_id')
        return ParticipantStage.objects.filter(stage_id=stage_id).select_related(
            'participant', 'stage', 'participant__assigned_bungalow'
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def stage_participants_stats(request, stage_id):
    """Retourne les statistiques des participants d'un événement."""
    try:
        stage = Stage.objects.get(pk=stage_id)
    except Stage.DoesNotExist:
        return Response(
            {'error': f'Événement non trouvé (ID: {stage_id})'},
            status=status.HTTP_404_NOT_FOUND
        )

    registrations = ParticipantStage.objects.filter(stage=stage)

    # Statistiques par rôle
    role_stats = {}
    for role_code, role_label in ParticipantStage.ROLE_CHOICES:
        role_stats[role_code] = registrations.filter(role=role_code).count()

    # Participants assignés à un bungalow
    assigned_count = registrations.filter(
        participant__assigned_bungalow__isnull=False
    ).count()

    return Response({
        'stageId': stage.id,
        'stageName': stage.name,
        'totalParticipants': registrations.count(),
        'capacity': stage.capacity,
        'availableSpots': max(0, stage.capacity - registrations.count()),
        'assignedToBungalow': assigned_count,
        'notAssigned': registrations.count() - assigned_count,
        'byRole': role_stats
    })


# ==================== PARTICIPANT SIMPLE VIEWS (sans événement) ====================

class ParticipantSimpleListCreateView(generics.ListCreateAPIView):
    """Vue pour lister et créer des participants (sans lien événement)."""

    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name', 'last_name', 'email', 'nationality']
    filterset_fields = ['gender', 'status']
    ordering_fields = ['last_name', 'first_name', 'created_at']
    ordering = ['last_name', 'first_name']

    def get_queryset(self):
        """Retourne tous les participants."""
        return Participant.objects.prefetch_related('languages', 'stage_participations').all()

    def get_serializer_class(self):
        """Retourne le serializer approprié selon la méthode HTTP."""
        if self.request.method == 'POST':
            return ParticipantCreateSimpleSerializer
        return ParticipantSimpleSerializer

    def perform_create(self, serializer):
        """Ajoute l'utilisateur créateur."""
        serializer.save(created_by=self.request.user)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_participants(request):
    """Recherche de participants pour l'ajout à un événement."""
    query = request.query_params.get('q', '')
    stage_id = request.query_params.get('exclude_stage', None)

    participants = Participant.objects.all()

    # Recherche par nom, prénom ou email
    if query:
        participants = participants.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )

    # Exclure les participants déjà inscrits à un événement spécifique
    if stage_id:
        already_registered = ParticipantStage.objects.filter(
            stage_id=stage_id
        ).values_list('participant_id', flat=True)
        participants = participants.exclude(id__in=already_registered)

    # Limiter les résultats
    participants = participants[:20]

    serializer = ParticipantSimpleSerializer(participants, many=True)
    return Response(serializer.data)


# ==================== PARTICIPANT STAGE ASSIGNMENT ====================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def unassigned_registrations(request):
    """
    Retourne toutes les inscriptions (ParticipantStage) non assignées à un bungalow.
    Un participant inscrit à 2 événements sans assignation retournera 2 entrées.
    """
    # Filtrer par dates si fournies
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')

    registrations = ParticipantStage.objects.filter(
        assigned_bungalow__isnull=True
    ).select_related('participant', 'stage', 'assigned_bungalow')

    # Filtrer par période (utiliser les dates effectives)
    if start_date and end_date:
        from datetime import datetime
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Filtrer les inscriptions dont la période chevauche la période demandée
            # Logique de chevauchement: le participant doit être présent pendant la période filtrée
            # - Si arrival_date et departure_date sont définis: utiliser ces dates
            # - Sinon: utiliser les dates du stage

            filtered_registrations = []
            for reg in registrations:
                # Déterminer les dates effectives du participant
                participant_start = reg.arrival_date if reg.arrival_date else reg.stage.start_date
                participant_end = reg.departure_date if reg.departure_date else reg.stage.end_date

                # Vérifier le chevauchement: le participant est présent si
                # sa date de départ est >= date de début du filtre ET
                # sa date d'arrivée est <= date de fin du filtre
                if participant_end >= start and participant_start <= end:
                    filtered_registrations.append(reg.id)

            registrations = registrations.filter(id__in=filtered_registrations)
        except ValueError:
            pass

    serializer = ParticipantStageSerializer(registrations, many=True)

    return Response({
        'registrations': serializer.data,
        'count': registrations.count()
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def assign_registration(request, registration_id):
    """
    Assigne une inscription (ParticipantStage) à un bungalow.
    La durée d'assignation est basée sur arrival_date/departure_date de l'inscription.
    """
    try:
        registration = ParticipantStage.objects.select_related('participant', 'stage').get(pk=registration_id)
    except ParticipantStage.DoesNotExist:
        return Response(
            {'error': f'Inscription non trouvée (ID: {registration_id})'},
            status=status.HTTP_404_NOT_FOUND
        )

    bungalow_id = request.data.get('bungalowId')
    bed_id = request.data.get('bed')

    if not bungalow_id:
        return Response(
            {'error': 'ID du bungalow manquant'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not bed_id:
        return Response(
            {'error': 'ID du lit manquant'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        bungalow = Bungalow.objects.get(pk=bungalow_id)
    except Bungalow.DoesNotExist:
        return Response(
            {'error': f'Bungalow non trouvé (ID: {bungalow_id})'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Vérifier que le lit existe dans le bungalow
    bed_exists = any(bed.get('id') == bed_id for bed in bungalow.beds)
    if not bed_exists:
        return Response(
            {'error': f'Le lit "{bed_id}" n\'existe pas dans le bungalow {bungalow.name}'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Dates effectives de l'inscription
    start_date = registration.effective_arrival_date
    end_date = registration.effective_departure_date

    # Vérifier si l'utilisateur veut forcer l'assignation malgré les règles
    force_assign = request.data.get('force_assign', False)

    # Vérifier les conflits de genre
    participant = registration.participant
    other_occupants = ParticipantStage.objects.filter(
        assigned_bungalow=bungalow
    ).exclude(id=registration_id).select_related('participant')

    for occupant in other_occupants:
        # Vérifier chevauchement de dates
        occ_start = occupant.effective_arrival_date
        occ_end = occupant.effective_departure_date
        if start_date <= occ_end and end_date >= occ_start:
            # Il y a chevauchement, vérifier le genre
            if occupant.participant.gender != participant.gender:
                warning_msg = (f'Conflit de genre: {occupant.participant.full_name} ({occupant.participant.get_gender_display()}) '
                             f'occupe ce bungalow du {occ_start} au {occ_end}. '
                             f'Voulez-vous vraiment ajouter {participant.full_name} ({participant.get_gender_display()}) ?')
                if not force_assign:
                    return Response({
                        'warning': True,
                        'message': warning_msg,
                        'requires_confirmation': True
                    }, status=status.HTTP_409_CONFLICT)

    # Vérifier si le lit est déjà occupé pendant cette période (BLOCAGE ABSOLU)
    bed_occupants = ParticipantStage.objects.filter(
        assigned_bungalow=bungalow,
        assigned_bed=bed_id
    ).exclude(id=registration_id)

    for occupant in bed_occupants:
        occ_start = occupant.effective_arrival_date
        occ_end = occupant.effective_departure_date
        if start_date <= occ_end and end_date >= occ_start:
            return Response({
                'error': f'Le lit {bed_id} est déjà occupé par {occupant.participant.full_name} '
                         f'du {occ_start} au {occ_end}. Impossible d\'assigner ce lit.'
            }, status=status.HTTP_400_BAD_REQUEST)

    # Vérifier si des participants de stages différents partagent le même bungalow pendant la même période
    other_stage_occupants = ParticipantStage.objects.filter(
        assigned_bungalow=bungalow
    ).exclude(id=registration_id).exclude(stage_id=registration.stage_id).select_related('participant', 'stage')

    for occupant in other_stage_occupants:
        occ_start = occupant.effective_arrival_date
        occ_end = occupant.effective_departure_date
        if start_date <= occ_end and end_date >= occ_start:
            warning_msg = (f'Conflit d\'événement: {occupant.participant.full_name} de l\'événement "{occupant.stage.name}" '
                         f'occupe ce bungalow du {occ_start} au {occ_end}. '
                         f'Voulez-vous vraiment ajouter {participant.full_name} de l\'événement "{registration.stage.name}" ?')
            if not force_assign:
                return Response({
                    'warning': True,
                    'message': warning_msg,
                    'requires_confirmation': True
                }, status=status.HTTP_409_CONFLICT)

    # ========== RÈGLES D'ASSIGNATION (MÊMES QUE L'AUTO-ASSIGNATION) ==========

    # Collecter les warnings au lieu de bloquer immédiatement
    warnings = []

    # RÈGLE 1: Les encadrants doivent être seuls dans leur chambre
    if registration.role == 'instructor':
        # Vérifier si d'autres personnes sont déjà assignées pendant la même période
        overlapping_occupants = ParticipantStage.objects.filter(
            assigned_bungalow=bungalow
        ).exclude(id=registration_id).select_related('participant')

        for occupant in overlapping_occupants:
            occ_start = occupant.effective_arrival_date
            occ_end = occupant.effective_departure_date
            if start_date <= occ_end and end_date >= occ_start:
                warning_msg = (f'Règle encadrants: Les encadrants doivent être seuls dans leur chambre. '
                             f'{occupant.participant.full_name} occupe déjà ce bungalow du {occ_start} au {occ_end}.')
                if not force_assign:
                    return Response({
                        'warning': True,
                        'message': warning_msg,
                        'requires_confirmation': True
                    }, status=status.HTTP_409_CONFLICT)
                else:
                    warnings.append(warning_msg)

    # RÈGLE 2: Ne pas assigner quelqu'un d'autre avec un encadrant
    overlapping_instructors = ParticipantStage.objects.filter(
        assigned_bungalow=bungalow,
        role='instructor'
    ).exclude(id=registration_id).select_related('participant')

    for instructor in overlapping_instructors:
        instr_start = instructor.effective_arrival_date
        instr_end = instructor.effective_departure_date
        if start_date <= instr_end and end_date >= instr_start:
            warning_msg = (f'Règle encadrants: Impossible d\'assigner à ce bungalow. '
                         f'L\'encadrant {instructor.participant.full_name} doit être seul et occupe ce bungalow '
                         f'du {instr_start} au {instr_end}.')
            if not force_assign:
                return Response({
                    'warning': True,
                    'message': warning_msg,
                    'requires_confirmation': True
                }, status=status.HTTP_409_CONFLICT)
            else:
                warnings.append(warning_msg)

    # RÈGLE 3: Les musiciens devraient être dans le Village C
    if registration.role == 'musician' and bungalow.village != 'C':
        warning_msg = (f'Règle musiciens: Les musiciens doivent être assignés au Village C. '
                     f'Le bungalow {bungalow.name} est dans le Village {bungalow.village}.')
        if not force_assign:
            return Response({
                'warning': True,
                'message': warning_msg,
                'requires_confirmation': True
            }, status=status.HTTP_409_CONFLICT)
        else:
            warnings.append(warning_msg)

    # RÈGLE 4: Ne pas mélanger les rôles (étudiants séparés des musiciens/encadrants)
    if registration.role == 'participant':
        # Vérifier qu'il n'y a pas de musiciens ou encadrants pendant la même période
        overlapping_others = ParticipantStage.objects.filter(
            assigned_bungalow=bungalow,
            role__in=['musician', 'instructor']
        ).exclude(id=registration_id).select_related('participant')

        for occupant in overlapping_others:
            occ_start = occupant.effective_arrival_date
            occ_end = occupant.effective_departure_date
            if start_date <= occ_end and end_date >= occ_start:
                role_display = 'musicien' if occupant.role == 'musician' else 'encadrant'
                warning_msg = (f'Règle séparation: Les étudiants ne peuvent pas partager un bungalow avec des musiciens ou encadrants. '
                             f'{occupant.participant.full_name} ({role_display}) occupe ce bungalow du {occ_start} au {occ_end}.')
                if not force_assign:
                    return Response({
                        'warning': True,
                        'message': warning_msg,
                        'requires_confirmation': True
                    }, status=status.HTTP_409_CONFLICT)
                else:
                    warnings.append(warning_msg)

    elif registration.role in ['musician', 'staff']:
        # Les musiciens/staff ne doivent pas être avec des étudiants
        overlapping_students = ParticipantStage.objects.filter(
            assigned_bungalow=bungalow,
            role='participant'
        ).exclude(id=registration_id).select_related('participant')

        for student in overlapping_students:
            std_start = student.effective_arrival_date
            std_end = student.effective_departure_date
            if start_date <= std_end and end_date >= std_start:
                role_display = 'musicien' if registration.role == 'musician' else 'staff'
                warning_msg = (f'Règle séparation: Les {role_display}s ne peuvent pas partager un bungalow avec des étudiants. '
                             f'{student.participant.full_name} (étudiant) occupe ce bungalow du {std_start} au {std_end}.')
                if not force_assign:
                    return Response({
                        'warning': True,
                        'message': warning_msg,
                        'requires_confirmation': True
                    }, status=status.HTTP_409_CONFLICT)
                else:
                    warnings.append(warning_msg)

    # ========== FIN DES RÈGLES D'ASSIGNATION ==========

    # Effectuer l'assignation dans une transaction atomique
    with transaction.atomic():
        registration.assigned_bungalow = bungalow
        registration.assigned_bed = bed_id
        registration.was_forced = force_assign  # Marquer si l'assignation a été forcée
        registration.save()

        # Mettre à jour les lits du bungalow
        for bed in bungalow.beds:
            if bed.get('id') == bed_id:
                bed['occupiedBy'] = {
                    'registrationId': registration.id,
                    'participantId': participant.id,
                    'name': participant.full_name,
                    'gender': participant.gender,
                    'age': participant.age if participant.age else None,
                    'nationality': participant.nationality if participant.nationality else None,
                    'languages': [lang.name for lang in participant.languages.all()] if participant.languages.exists() else [],
                    'role': registration.role,
                    'startDate': str(start_date),
                    'startTime': str(registration.arrival_time) if registration.arrival_time else '',
                    'endDate': str(end_date),
                    'endTime': str(registration.departure_time) if registration.departure_time else '',
                    'stageName': registration.stage.name,
                    'wasForced': registration.was_forced  # Indiquer si l'assignation a été forcée
                }
                break
        bungalow.save()
        bungalow.update_occupancy()

        # Log de l'activité
        log_assignment(request.user, participant, bungalow, bed_id)

    # Recharger l'inscription pour s'assurer d'avoir les données à jour
    registration.refresh_from_db()
    serializer = ParticipantStageSerializer(registration)
    return Response({
        'success': True,
        'message': f'{participant.full_name} assigné au bungalow {bungalow.name} (lit: {bed_id}) '
                   f'du {start_date} au {end_date}',
        'registration': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def unassign_registration(request, registration_id):
    """Désassigne une inscription de son bungalow."""
    try:
        registration = ParticipantStage.objects.select_related('participant', 'assigned_bungalow').get(pk=registration_id)
    except ParticipantStage.DoesNotExist:
        return Response(
            {'error': f'Inscription non trouvée (ID: {registration_id})'},
            status=status.HTTP_404_NOT_FOUND
        )

    if not registration.is_assigned:
        return Response(
            {'error': f'{registration.participant.full_name} n\'est pas assigné pour cet événement'},
            status=status.HTTP_400_BAD_REQUEST
        )

    old_bungalow = registration.assigned_bungalow
    old_bed = registration.assigned_bed

    # Libérer le lit dans le bungalow
    if old_bungalow:
        for bed in old_bungalow.beds:
            if bed.get('id') == old_bed:
                occupant = bed.get('occupiedBy')
                if occupant and occupant.get('registrationId') == registration.id:
                    bed['occupiedBy'] = None
                break
        old_bungalow.save()
        old_bungalow.update_occupancy()

    # Désassigner
    registration.assigned_bungalow = None
    registration.assigned_bed = None
    registration.save()

    # Log de l'activité
    log_unassignment(request.user, registration.participant, old_bungalow, old_bed)

    serializer = ParticipantStageSerializer(registration)
    return Response({
        'success': True,
        'message': f'{registration.participant.full_name} désassigné du bungalow {old_bungalow.name if old_bungalow else "inconnu"}',
        'registration': serializer.data
    })


# ==================== EXPORT ASSIGNATIONS ====================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_assignments(request):
    """
    Exporte les assignations au format JSON pour le frontend.
    Paramètres optionnels:
    - stage_id: filtrer par événement
    """
    stage_id = request.query_params.get('stage_id')

    # Récupérer toutes les inscriptions assignées
    queryset = ParticipantStage.objects.filter(
        assigned_bungalow__isnull=False
    ).select_related('participant', 'stage', 'assigned_bungalow', 'assigned_bungalow__village')

    # Filtrer par événement si spécifié
    if stage_id:
        queryset = queryset.filter(stage_id=stage_id)

    # Construire les données d'export
    assignments = []
    for reg in queryset:
        assignments.append({
            'village': f"Village {reg.assigned_bungalow.village.name}",
            'bungalow': reg.assigned_bungalow.name,
            'participantName': reg.participant.full_name,
            'arrivalDate': str(reg.effective_arrival_date) if reg.effective_arrival_date else '',
            'arrivalTime': reg.arrival_time or '',
            'departureDate': str(reg.effective_departure_date) if reg.effective_departure_date else '',
            'departureTime': reg.departure_time or '',
            'stageName': reg.stage.name
        })

    return Response({
        'assignments': assignments,
        'count': len(assignments)
    })


# ==================== EXCEL IMPORT ====================

from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.decorators import parser_classes
from datetime import datetime, date, time


def parse_excel_date(value):
    """
    Parse une date depuis une cellule Excel.
    Retourne (date_string, error_message) - si error_message est None, c'est valide.
    """
    if value is None or value == '':
        return '', None

    # Si c'est déjà un objet date/datetime
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d'), None
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d'), None

    # Si c'est une string, la nettoyer et valider
    value_str = str(value).strip().replace('\xa0', ' ').strip()
    if not value_str:
        return '', None

    # Essayer différents formats de date
    date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']
    for fmt in date_formats:
        try:
            parsed = datetime.strptime(value_str, fmt)
            return parsed.strftime('%Y-%m-%d'), None
        except ValueError:
            continue

    # Si ça contient une heure (datetime string), extraire juste la date
    try:
        # Format: "2025-12-03 00:00:00"
        if ' ' in value_str:
            date_part = value_str.split(' ')[0]
            for fmt in date_formats:
                try:
                    parsed = datetime.strptime(date_part, fmt)
                    return parsed.strftime('%Y-%m-%d'), None
                except ValueError:
                    continue
    except:
        pass

    return None, f'Format de date invalide: "{value_str}". Utilisez AAAA-MM-JJ (ex: 2025-12-03)'


def parse_excel_time(value):
    """
    Parse une heure depuis une cellule Excel.
    Retourne (time_string, error_message) - si error_message est None, c'est valide.
    """
    if value is None or value == '':
        return '', None

    # Si c'est déjà un objet time/datetime
    if isinstance(value, time):
        return value.strftime('%H:%M'), None
    if isinstance(value, datetime):
        return value.strftime('%H:%M'), None

    # Si c'est une string, la nettoyer et valider
    value_str = str(value).strip().replace('\xa0', ' ').strip()
    if not value_str:
        return '', None

    # Essayer différents formats d'heure
    time_formats = ['%H:%M', '%H:%M:%S', '%Hh%M', '%H h %M']
    for fmt in time_formats:
        try:
            parsed = datetime.strptime(value_str, fmt)
            return parsed.strftime('%H:%M'), None
        except ValueError:
            continue

    return None, f'Format d\'heure invalide: "{value_str}". Utilisez HH:MM (ex: 14:30)'


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def validate_excel_import(request):
    """
    Valide un fichier Excel et retourne les données à importer.
    Identifie les participants existants, non-existants, et les erreurs.

    Format Excel attendu:
    - email (obligatoire, identifiant unique)
    - first_name (obligatoire si nouveau participant)
    - last_name (obligatoire si nouveau participant)
    - stage_name (obligatoire)
    - role (optionnel: participant, musician, instructor, staff)
    - arrival_date (optionnel: YYYY-MM-DD)
    - arrival_time (optionnel: HH:MM)
    - departure_date (optionnel: YYYY-MM-DD)
    - departure_time (optionnel: HH:MM)
    - gender (optionnel si nouveau: M/F)
    - age (optionnel si nouveau)
    - nationality (optionnel si nouveau)
    - status (optionnel si nouveau: student, instructor, professional, staff)
    - languages (optionnel: noms des langues séparés par virgule, ex: "Français, English, Wolof")
    """
    import openpyxl
    from io import BytesIO

    import csv

    # Vérifier que le fichier est présent
    if 'file' not in request.FILES:
        return Response({'error': 'Aucun fichier fourni'}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES['file']
    filename = file.name.lower()

    # Vérifier l'extension
    if not filename.endswith(('.xlsx', '.xls', '.csv')):
        return Response({'error': 'Format de fichier non supporté. Utilisez .xlsx, .xls ou .csv'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = []

        if filename.endswith('.csv'):
            # Lire le fichier CSV - essayer plusieurs encodages
            file_bytes = file.read()
            file_content = None

            # Liste des encodages à essayer (du plus courant au moins courant)
            encodings_to_try = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']

            for encoding in encodings_to_try:
                try:
                    file_content = file_bytes.decode(encoding)
                    break
                except (UnicodeDecodeError, LookupError):
                    continue

            if file_content is None:
                return Response({
                    'error': 'Impossible de lire le fichier. Veuillez le sauvegarder en UTF-8 ou utiliser le format Excel (.xlsx).'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Détecter le délimiteur (tabulation, point-virgule ou virgule)
            first_line = file_content.split('\n')[0] if '\n' in file_content else file_content

            # Ordre de priorité: tabulation > point-virgule > virgule
            if '\t' in first_line:
                delimiter = '\t'
            elif ';' in first_line:
                delimiter = ';'
            else:
                delimiter = ','

            reader = csv.reader(file_content.splitlines(), delimiter=delimiter)
            rows = list(reader)

            if len(rows) < 1:
                return Response({'error': 'Le fichier CSV est vide'}, status=status.HTTP_400_BAD_REQUEST)

            # Récupérer les en-têtes (première ligne) et nettoyer les caractères invisibles
            headers = [h.lower().strip().replace('\ufeff', '').replace('\u200b', '') for h in rows[0]]
            data_rows = rows[1:]

            # Debug: si aucune colonne reconnue, montrer ce qu'on a trouvé
            if 'email' not in headers and 'stage_name' not in headers:
                return Response({
                    'error': f'Format de fichier non reconnu. Colonnes détectées: {", ".join(headers[:5])}... Veuillez utiliser le template fourni.'
                }, status=status.HTTP_400_BAD_REQUEST)
        else:
            # Lire le fichier Excel
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active

            # Récupérer les en-têtes (première ligne)
            headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))

        # Vérifier les colonnes obligatoires
        required_columns = ['email', 'stage_name']
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            return Response({
                'error': f'Colonnes manquantes: {", ".join(missing_columns)}'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Récupérer tous les stages existants
        stages_by_name = {s.name.lower(): s for s in Stage.objects.all()}

        # Récupérer tous les participants existants par email
        participants_by_email = {p.email.lower(): p for p in Participant.objects.all()}

        # Récupérer toutes les langues existantes (par nom, insensible à la casse)
        languages_by_name = {l.name.lower(): l for l in Language.objects.filter(is_active=True)}

        results = {
            'valid_imports': [],           # Participants existants à ajouter
            'new_participants': [],         # Participants à créer
            'errors': [],                   # Erreurs non récupérables
            'already_registered': [],       # Déjà inscrits à l'événement
        }

        row_num = 1
        for row in data_rows:
            row_num += 1
            row_data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}

            email = str(row_data.get('email', '')).strip().lower() if row_data.get('email') else ''
            stage_name = str(row_data.get('stage_name', '')).strip() if row_data.get('stage_name') else ''

            # Ignorer les lignes vides
            if not email and not stage_name:
                continue

            # Vérifier email
            if not email:
                results['errors'].append({
                    'row': row_num,
                    'data': row_data,
                    'reason': 'Email manquant'
                })
                continue

            # Vérifier si l'événement existe
            if not stage_name:
                results['errors'].append({
                    'row': row_num,
                    'email': email,
                    'data': row_data,
                    'reason': 'Nom d\'événement manquant'
                })
                continue

            stage = stages_by_name.get(stage_name.lower())
            if not stage:
                results['errors'].append({
                    'row': row_num,
                    'email': email,
                    'stageName': stage_name,
                    'data': row_data,
                    'reason': f'Événement "{stage_name}" non trouvé'
                })
                continue

            # Parser et valider les dates/heures
            arrival_date, arrival_date_error = parse_excel_date(row_data.get('arrival_date'))
            arrival_time, arrival_time_error = parse_excel_time(row_data.get('arrival_time'))
            departure_date, departure_date_error = parse_excel_date(row_data.get('departure_date'))
            departure_time, departure_time_error = parse_excel_time(row_data.get('departure_time'))

            # Collecter les erreurs de date/heure
            date_errors = []
            if arrival_date_error:
                date_errors.append(f"Date d'arrivée: {arrival_date_error}")
            if arrival_time_error:
                date_errors.append(f"Heure d'arrivée: {arrival_time_error}")
            if departure_date_error:
                date_errors.append(f"Date de départ: {departure_date_error}")
            if departure_time_error:
                date_errors.append(f"Heure de départ: {departure_time_error}")

            if date_errors:
                results['errors'].append({
                    'row': row_num,
                    'email': email,
                    'stageName': stage.name,
                    'data': row_data,
                    'reason': ' | '.join(date_errors)
                })
                continue

            # Parser les langues (séparées par virgule)
            languages_str = str(row_data.get('languages', '')).strip() if row_data.get('languages') else ''
            language_ids = []
            language_names = []
            unknown_languages = []

            if languages_str:
                for lang_name in languages_str.split(','):
                    lang_name = lang_name.strip()
                    if lang_name:
                        lang = languages_by_name.get(lang_name.lower())
                        if lang:
                            language_ids.append(lang.id)
                            language_names.append(lang.name)
                        else:
                            unknown_languages.append(lang_name)

            # Avertir si des langues n'existent pas (mais ne pas bloquer)
            language_warning = None
            if unknown_languages:
                language_warning = f"Langues non trouvées: {', '.join(unknown_languages)}"

            # Préparer les données d'inscription
            registration_data = {
                'stageId': stage.id,
                'stageName': stage.name,
                'role': row_data.get('role', 'participant') or 'participant',
                'arrivalDate': arrival_date,
                'arrivalTime': arrival_time,
                'departureDate': departure_date,
                'departureTime': departure_time,
                'languageIds': language_ids,
                'languageNames': language_names,
                'languageWarning': language_warning,
            }

            # Vérifier si le participant existe
            participant = participants_by_email.get(email)

            if participant:
                # Vérifier si déjà inscrit à cet événement
                already_registered = ParticipantStage.objects.filter(
                    participant=participant,
                    stage=stage
                ).exists()

                if already_registered:
                    results['already_registered'].append({
                        'row': row_num,
                        'email': email,
                        'participantId': participant.id,
                        'participantName': f'{participant.first_name} {participant.last_name}',
                        'stageName': stage.name,
                        'reason': 'Déjà inscrit à cet événement'
                    })
                else:
                    results['valid_imports'].append({
                        'row': row_num,
                        'email': email,
                        'participantId': participant.id,
                        'participantName': f'{participant.first_name} {participant.last_name}',
                        **registration_data
                    })
            else:
                # Nouveau participant - récupérer les infos pour création
                first_name = str(row_data.get('first_name', '')).strip() if row_data.get('first_name') else ''
                last_name = str(row_data.get('last_name', '')).strip() if row_data.get('last_name') else ''

                if not first_name or not last_name:
                    results['errors'].append({
                        'row': row_num,
                        'email': email,
                        'data': row_data,
                        'reason': 'Prénom et nom requis pour un nouveau participant'
                    })
                    continue

                results['new_participants'].append({
                    'row': row_num,
                    'email': email,
                    'firstName': first_name,
                    'lastName': last_name,
                    'gender': row_data.get('gender', 'F') or 'F',
                    'age': int(row_data.get('age', 25)) if row_data.get('age') else 25,
                    'nationality': str(row_data.get('nationality', '')).strip() if row_data.get('nationality') else '',
                    'status': row_data.get('status', 'student') or 'student',
                    **registration_data
                })

        return Response({
            'summary': {
                'totalRows': row_num - 1,
                'validImports': len(results['valid_imports']),
                'newParticipants': len(results['new_participants']),
                'alreadyRegistered': len(results['already_registered']),
                'errors': len(results['errors'])
            },
            **results
        })

    except UnicodeDecodeError:
        return Response({
            'error': 'Le fichier contient des caractères non reconnus. Veuillez sauvegarder votre fichier en UTF-8 ou utiliser le format Excel (.xlsx).'
        }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        # Transformer les erreurs techniques en messages compréhensibles
        error_msg = str(e)
        if 'codec' in error_msg or 'decode' in error_msg or 'encode' in error_msg:
            user_message = 'Le fichier contient des caractères non reconnus. Veuillez sauvegarder votre fichier en UTF-8 ou utiliser le format Excel (.xlsx).'
        elif 'openpyxl' in error_msg or 'workbook' in error_msg.lower():
            user_message = 'Le fichier Excel semble corrompu ou dans un format non supporté. Veuillez utiliser le format .xlsx.'
        elif 'permission' in error_msg.lower():
            user_message = 'Impossible d\'accéder au fichier. Veuillez réessayer.'
        else:
            user_message = 'Une erreur est survenue lors de la lecture du fichier. Veuillez vérifier le format et réessayer.'
        return Response({
            'error': user_message
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def execute_excel_import(request):
    """
    Exécute l'import après validation.
    Reçoit les listes de participants à importer et à créer.
    """
    valid_imports = request.data.get('valid_imports', [])
    new_participants = request.data.get('new_participants', [])

    # Récupérer le nom du stage pour le log
    stage_name = None
    if valid_imports:
        stage_name = valid_imports[0].get('stageName')
    elif new_participants:
        stage_name = new_participants[0].get('stageName')

    results = {
        'imported': [],
        'created_and_imported': [],
        'errors': []
    }

    # Importer les participants existants
    for item in valid_imports:
        try:
            participant_stage = ParticipantStage.objects.create(
                participant_id=item['participantId'],
                stage_id=item['stageId'],
                role=item.get('role', 'participant'),
                arrival_date=item.get('arrivalDate') or None,
                arrival_time=item.get('arrivalTime') or None,
                departure_date=item.get('departureDate') or None,
                departure_time=item.get('departureTime') or None,
                created_by=request.user
            )

            # Ajouter les langues au participant existant si spécifiées
            language_ids = item.get('languageIds', [])
            if language_ids:
                participant = Participant.objects.get(id=item['participantId'])
                # Ajouter les nouvelles langues sans supprimer les existantes
                participant.languages.add(*language_ids)

            # Mettre à jour le compteur de participants (recalculer avec filtre role='participant')
            stage = Stage.objects.get(id=item['stageId'])
            stage.current_participants = stage.participant_registrations.filter(role='participant').count()
            stage.save(update_fields=['current_participants'])

            # Log individuel pour ce participant
            log_excel_import_participant(
                user=request.user,
                participant_name=item['participantName'],
                stage_name=item['stageName'],
                was_created=False,
                languages=item.get('languageNames', [])
            )

            results['imported'].append({
                'email': item['email'],
                'participantName': item['participantName'],
                'stageName': item['stageName'],
                'languagesAdded': item.get('languageNames', [])
            })
        except Exception as e:
            results['errors'].append({
                'email': item['email'],
                'reason': str(e)
            })

    # Créer et importer les nouveaux participants
    for item in new_participants:
        try:
            # Créer le participant
            participant = Participant.objects.create(
                first_name=item['firstName'],
                last_name=item['lastName'],
                email=item['email'],
                gender=item.get('gender', 'F'),
                age=item.get('age', 25),
                nationality=item.get('nationality', ''),
                status=item.get('status', 'student'),
                created_by=request.user
            )

            # Ajouter les langues au nouveau participant
            language_ids = item.get('languageIds', [])
            if language_ids:
                participant.languages.add(*language_ids)

            # L'ajouter à l'événement
            ParticipantStage.objects.create(
                participant=participant,
                stage_id=item['stageId'],
                role=item.get('role', 'participant'),
                arrival_date=item.get('arrivalDate') or None,
                arrival_time=item.get('arrivalTime') or None,
                departure_date=item.get('departureDate') or None,
                departure_time=item.get('departureTime') or None,
                created_by=request.user
            )

            # Mettre à jour le compteur de participants (recalculer avec filtre role='participant')
            stage = Stage.objects.get(id=item['stageId'])
            stage.current_participants = stage.participant_registrations.filter(role='participant').count()
            stage.save(update_fields=['current_participants'])

            # Log individuel pour ce nouveau participant
            participant_name = f"{item['firstName']} {item['lastName']}"
            log_excel_import_participant(
                user=request.user,
                participant_name=participant_name,
                stage_name=item['stageName'],
                was_created=True,
                languages=item.get('languageNames', [])
            )

            results['created_and_imported'].append({
                'email': item['email'],
                'participantName': participant_name,
                'stageName': item['stageName'],
                'languagesAdded': item.get('languageNames', [])
            })
        except Exception as e:
            results['errors'].append({
                'email': item['email'],
                'reason': str(e)
            })

    # Log résumé de l'import Excel (si au moins un participant a été importé)
    imported_count = len(results['imported'])
    created_count = len(results['created_and_imported'])
    if (imported_count > 0 or created_count > 0) and stage_name:
        log_excel_import_summary(request.user, stage_name, imported_count, created_count)

    return Response({
        'summary': {
            'imported': len(results['imported']),
            'createdAndImported': len(results['created_and_imported']),
            'errors': len(results['errors'])
        },
        **results
    })


# ==================== BILAN DE FREQUENTATION ====================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def frequency_report(request):
    """
    Génère un bilan de fréquentation pour une période donnée.

    Paramètres:
    - start_date: Date de début (YYYY-MM-DD)
    - end_date: Date de fin (YYYY-MM-DD)

    Retourne:
    - Nombre de stages, résidences, autres activités
    - Nombre d'enseignants, élèves, autres catégories
    - Répartition hommes/femmes
    - Âge moyen
    - Nationalités et leur répartition
    - Taux de fréquentation (capacité vs remplissage)
    """
    from django.db.models import Count, Avg, Min, Max, Sum
    from collections import Counter

    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')

    if not start_date or not end_date:
        return Response(
            {'error': 'Les paramètres start_date et end_date sont requis (format: YYYY-MM-DD)'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        from datetime import datetime
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        return Response(
            {'error': 'Format de date invalide. Utilisez YYYY-MM-DD'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # ========== ÉVÉNEMENTS ==========
    # Filtrer les événements qui se chevauchent avec la période
    events = Stage.objects.filter(
        Q(start_date__lte=end, end_date__gte=start)
    )

    # Compter par type d'événement
    events_by_type = events.values('event_type').annotate(count=Count('id'))
    event_counts = {item['event_type']: item['count'] for item in events_by_type}

    stages_count = event_counts.get('stage', 0)
    residents_count = event_counts.get('resident', 0)
    autres_count = event_counts.get('autres', 0)
    total_events = events.count()

    # Liste des événements pour référence (avec le nombre réel de participants role='participant')
    events_list = []
    for e in events:
        # Compter le nombre réel de participants (role='participant' uniquement)
        real_participant_count = ParticipantStage.objects.filter(stage=e, role='participant').count()
        events_list.append({
            'id': e.id,
            'name': e.name,
            'type': e.event_type,
            'startDate': str(e.start_date),
            'endDate': str(e.end_date),
            'capacity': e.capacity,
            'currentParticipants': real_participant_count
        })

    # ========== INSCRIPTIONS DANS LA PÉRIODE ==========
    # Récupérer toutes les inscriptions pour les événements de la période
    registrations = ParticipantStage.objects.filter(
        stage__in=events
    ).select_related('participant', 'stage')

    # IDs des participants UNIQUES (une personne peut participer à plusieurs événements)
    unique_participant_ids = registrations.values_list('participant_id', flat=True).distinct()
    unique_participants = Participant.objects.filter(id__in=unique_participant_ids)

    total_registrations = registrations.count()
    unique_participants_count = unique_participants.count()

    # ========== STATISTIQUES PAR RÔLE ==========
    role_stats = registrations.values('role').annotate(count=Count('id'))
    role_counts = {item['role']: item['count'] for item in role_stats}

    instructors_count = role_counts.get('instructor', 0)
    participants_count = role_counts.get('participant', 0)
    musicians_count = role_counts.get('musician', 0)
    staff_count = role_counts.get('staff', 0)

    # ========== STATISTIQUES PAR STATUT (basé sur les participants uniques) ==========
    status_stats = unique_participants.values('status').annotate(count=Count('id'))
    status_counts = {item['status']: item['count'] for item in status_stats}

    students_count = status_counts.get('student', 0)
    instructors_status_count = status_counts.get('instructor', 0)
    professionals_count = status_counts.get('professional', 0)
    staff_status_count = status_counts.get('staff', 0)

    # ========== RÉPARTITION PAR GENRE ==========
    gender_stats = unique_participants.values('gender').annotate(count=Count('id'))
    gender_counts = {item['gender']: item['count'] for item in gender_stats}

    men_count = gender_counts.get('M', 0)
    women_count = gender_counts.get('F', 0)

    # ========== ÂGE MOYEN ==========
    age_data = unique_participants.aggregate(
        avg_age=Avg('age'),
        min_age=Min('age'),
        max_age=Max('age')
    )
    average_age = round(age_data['avg_age'], 1) if age_data['avg_age'] else 0
    min_age = age_data['min_age'] or 0
    max_age = age_data['max_age'] or 0

    # Distribution des âges par tranche
    age_ranges = [
        ('0-17', 0, 17),
        ('18-25', 18, 25),
        ('26-35', 26, 35),
        ('36-45', 36, 45),
        ('46-55', 46, 55),
        ('56-65', 56, 65),
        ('66+', 66, 200)
    ]

    age_distribution = {}
    for label, min_a, max_a in age_ranges:
        count = unique_participants.filter(age__gte=min_a, age__lte=max_a).count()
        age_distribution[label] = count

    # ========== NATIONALITÉS ==========
    # Compter les nationalités des participants uniques
    nationality_data = unique_participants.exclude(
        nationality__isnull=True
    ).exclude(
        nationality=''
    ).values('nationality').annotate(count=Count('id')).order_by('-count')

    nationalities = [
        {'nationality': item['nationality'], 'count': item['count']}
        for item in nationality_data
    ]
    total_nationalities = len(nationalities)

    # ========== TAUX DE FRÉQUENTATION ==========
    # Capacité totale des bungalows
    total_bed_capacity = Bungalow.objects.aggregate(total=Sum('capacity'))['total'] or 0

    # Calculer l'occupation moyenne sur la période
    # Pour simplifier, on compte les inscriptions assignées à des bungalows
    assigned_registrations = registrations.filter(assigned_bungalow__isnull=False).count()

    # Capacité totale des événements
    total_event_capacity = events.aggregate(total=Sum('capacity'))['total'] or 0

    # Taux de remplissage des événements
    event_fill_rate = round((total_registrations / total_event_capacity * 100), 1) if total_event_capacity > 0 else 0

    # Taux d'assignation aux bungalows
    assignment_rate = round((assigned_registrations / total_registrations * 100), 1) if total_registrations > 0 else 0

    # ========== LANGUES ==========
    # Langues parlées par les participants uniques
    language_data = Language.objects.filter(
        participants__in=unique_participants
    ).annotate(
        count=Count('participants', filter=Q(participants__in=unique_participants))
    ).values('name', 'count').order_by('-count')

    languages = [
        {'language': item['name'], 'count': item['count']}
        for item in language_data
    ]

    # ========== CONSTRUIRE LA RÉPONSE ==========
    return Response({
        'period': {
            'startDate': start_date,
            'endDate': end_date
        },
        'events': {
            'total': total_events,
            'stages': stages_count,
            'residences': residents_count,
            'autres': autres_count,
            'list': events_list
        },
        'participants': {
            'totalRegistrations': total_registrations,
            'uniqueParticipants': unique_participants_count,
            'byRole': {
                'participants': participants_count,
                'instructors': instructors_count,
                'musicians': musicians_count,
                'staff': staff_count
            },
            'byStatus': {
                'students': students_count,
                'instructors': instructors_status_count,
                'professionals': professionals_count,
                'staff': staff_status_count
            }
        },
        'demographics': {
            'gender': {
                'men': men_count,
                'women': women_count,
                'menPercentage': round((men_count / unique_participants_count * 100), 1) if unique_participants_count > 0 else 0,
                'womenPercentage': round((women_count / unique_participants_count * 100), 1) if unique_participants_count > 0 else 0
            },
            'age': {
                'average': average_age,
                'min': min_age,
                'max': max_age,
                'distribution': age_distribution
            }
        },
        'nationalities': {
            'total': total_nationalities,
            'list': nationalities
        },
        'languages': languages,
        'occupancy': {
            'totalBedCapacity': total_bed_capacity,
            'totalEventCapacity': total_event_capacity,
            'totalRegistrations': total_registrations,
            'assignedToBungalows': assigned_registrations,
            'eventFillRate': event_fill_rate,
            'assignmentRate': assignment_rate
        }
    })


# ==================== DASHBOARD STATISTICS ====================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    """
    Génère les statistiques complètes pour le tableau de bord.

    Retourne:
    - Statistiques générales (événements, participants, bungalows)
    - Événements en cours et à venir
    - Taux d'occupation des villages
    - Activités récentes
    - Tendances et alertes
    """
    from django.db.models import Count, Avg, Sum, Min, Max
    from datetime import datetime, timedelta
    from collections import Counter

    today = timezone.now().date()

    # ========== ÉVÉNEMENTS ==========
    all_events = Stage.objects.all()

    # Événements actifs (en cours aujourd'hui)
    active_events = all_events.filter(start_date__lte=today, end_date__gte=today)

    # Événements à venir (dans les 30 prochains jours)
    upcoming_events = all_events.filter(
        start_date__gt=today,
        start_date__lte=today + timedelta(days=30)
    ).order_by('start_date')

    # Événements passés (terminés)
    past_events = all_events.filter(end_date__lt=today)

    # Événements par type
    events_by_type = all_events.values('event_type').annotate(count=Count('id'))
    event_type_counts = {item['event_type']: item['count'] for item in events_by_type}

    # Liste des événements actifs avec détails (role='participant' uniquement)
    active_events_list = []
    for event in active_events:
        registrations_count = ParticipantStage.objects.filter(stage=event, role='participant').count()
        assigned_count = ParticipantStage.objects.filter(stage=event, role='participant', assigned_bungalow__isnull=False).count()
        days_remaining = (event.end_date - today).days
        fill_rate = round((registrations_count / event.capacity * 100), 1) if event.capacity > 0 else 0

        active_events_list.append({
            'id': event.id,
            'name': event.name,
            'type': event.event_type,
            'startDate': str(event.start_date),
            'endDate': str(event.end_date),
            'capacity': event.capacity,
            'registrations': registrations_count,
            'assigned': assigned_count,
            'daysRemaining': days_remaining,
            'fillRate': fill_rate,
            'instructor': event.instructor
        })

    # Liste des événements à venir (role='participant' uniquement)
    upcoming_events_list = []
    for event in upcoming_events[:5]:  # Top 5
        registrations_count = ParticipantStage.objects.filter(stage=event, role='participant').count()
        days_until = (event.start_date - today).days
        fill_rate = round((registrations_count / event.capacity * 100), 1) if event.capacity > 0 else 0

        upcoming_events_list.append({
            'id': event.id,
            'name': event.name,
            'type': event.event_type,
            'startDate': str(event.start_date),
            'endDate': str(event.end_date),
            'capacity': event.capacity,
            'registrations': registrations_count,
            'daysUntil': days_until,
            'fillRate': fill_rate
        })

    # ========== PARTICIPANTS ==========
    all_participants = Participant.objects.all()
    total_participants = all_participants.count()

    # Participants par statut
    status_stats = all_participants.values('status').annotate(count=Count('id'))
    status_counts = {item['status']: item['count'] for item in status_stats}

    # Participants par genre
    gender_stats = all_participants.values('gender').annotate(count=Count('id'))
    gender_counts = {item['gender']: item['count'] for item in gender_stats}

    # Nouveaux participants ce mois
    first_day_of_month = today.replace(day=1)
    new_this_month = all_participants.filter(created_at__date__gte=first_day_of_month).count()

    # Âge moyen
    age_data = all_participants.aggregate(avg_age=Avg('age'))
    average_age = round(age_data['avg_age'], 1) if age_data['avg_age'] else 0

    # Top nationalités
    nationality_stats = all_participants.exclude(
        nationality__isnull=True
    ).exclude(nationality='').values('nationality').annotate(
        count=Count('id')
    ).order_by('-count')[:10]

    top_nationalities = [
        {'nationality': item['nationality'], 'count': item['count']}
        for item in nationality_stats
    ]

    # ========== INSCRIPTIONS ==========
    all_registrations = ParticipantStage.objects.all()

    # Inscriptions aux événements actifs
    active_registrations = ParticipantStage.objects.filter(stage__in=active_events)
    total_active_registrations = active_registrations.count()
    assigned_registrations = active_registrations.filter(assigned_bungalow__isnull=False).count()
    unassigned_registrations = active_registrations.filter(assigned_bungalow__isnull=True).count()

    # Inscriptions par rôle
    role_stats = active_registrations.values('role').annotate(count=Count('id'))
    role_counts = {item['role']: item['count'] for item in role_stats}

    # ========== BUNGALOWS ET VILLAGES ==========
    all_bungalows = Bungalow.objects.all()
    all_villages = Village.objects.all()
    total_bungalows = all_bungalows.count()
    total_bed_capacity = all_bungalows.aggregate(total=Sum('capacity'))['total'] or 0

    # Occupation actuelle (pour les événements en cours)
    active_assigned_bungalow_ids = active_registrations.filter(
        assigned_bungalow__isnull=False
    ).values_list('assigned_bungalow_id', flat=True).distinct()
    occupied_bungalows = len(set(active_assigned_bungalow_ids))

    # Statistiques par village
    village_stats = []
    for village in all_villages:
        village_bungalows = all_bungalows.filter(village=village)
        total_capacity = village_bungalows.aggregate(total=Sum('capacity'))['total'] or 0
        bungalow_count = village_bungalows.count()

        # Occupants actuels dans ce village
        current_occupants = active_registrations.filter(
            assigned_bungalow__village=village
        ).count()

        occupancy_rate = round((current_occupants / total_capacity * 100), 1) if total_capacity > 0 else 0

        village_stats.append({
            'id': village.id,
            'name': village.name,
            'bungalowCount': bungalow_count,
            'totalCapacity': total_capacity,
            'currentOccupants': current_occupants,
            'occupancyRate': occupancy_rate
        })

    # ========== ALERTES ET CONFLITS ==========
    alerts = []

    # Événements presque pleins (> 90%) - role='participant' uniquement
    for event in active_events:
        reg_count = ParticipantStage.objects.filter(stage=event, role='participant').count()
        if event.capacity > 0 and (reg_count / event.capacity) > 0.9:
            alerts.append({
                'type': 'capacity',
                'severity': 'warning',
                'message': f"L'événement '{event.name}' est presque plein ({reg_count}/{event.capacity})",
                'eventId': event.id
            })

    # Événements qui se terminent bientôt
    ending_soon = active_events.filter(end_date__lte=today + timedelta(days=3))
    for event in ending_soon:
        days_left = (event.end_date - today).days
        alerts.append({
            'type': 'ending',
            'severity': 'info',
            'message': f"'{event.name}' se termine dans {days_left} jour(s)",
            'eventId': event.id
        })

    # Inscriptions non assignées
    if unassigned_registrations > 0:
        alerts.append({
            'type': 'unassigned',
            'severity': 'warning',
            'message': f"{unassigned_registrations} inscription(s) non assignée(s) à un logement"
        })

    # ========== ACTIVITÉS RÉCENTES ==========
    recent_activities = ActivityLog.objects.select_related('user').all().order_by('-timestamp')[:10]
    activities_list = []
    for activity in recent_activities:
        activities_list.append({
            'id': activity.id,
            'actionType': activity.action_type,
            'entityType': activity.model_name,
            'entityName': activity.object_repr,
            'description': activity.description,
            'timestamp': activity.timestamp.isoformat(),
            'user': activity.user.get_full_name() or activity.user.username if activity.user else 'Système'
        })

    # ========== TENDANCES (comparaison avec le mois précédent) ==========
    last_month_start = (first_day_of_month - timedelta(days=1)).replace(day=1)
    last_month_end = first_day_of_month - timedelta(days=1)

    # Nouveaux participants le mois dernier
    new_last_month = Participant.objects.filter(
        created_at__date__gte=last_month_start,
        created_at__date__lte=last_month_end
    ).count()

    # Événements ce mois vs mois dernier
    events_this_month = Stage.objects.filter(
        start_date__gte=first_day_of_month,
        start_date__lte=today
    ).count()

    events_last_month = Stage.objects.filter(
        start_date__gte=last_month_start,
        start_date__lte=last_month_end
    ).count()

    # Calculer les tendances
    participant_trend = ((new_this_month - new_last_month) / new_last_month * 100) if new_last_month > 0 else (100 if new_this_month > 0 else 0)
    event_trend = ((events_this_month - events_last_month) / events_last_month * 100) if events_last_month > 0 else (100 if events_this_month > 0 else 0)

    # ========== LANGUES LES PLUS PARLÉES ==========
    language_stats = Language.objects.annotate(
        speaker_count=Count('participants')
    ).filter(speaker_count__gt=0).order_by('-speaker_count')[:5]

    top_languages = [
        {'name': lang.name, 'code': lang.code, 'count': lang.speaker_count}
        for lang in language_stats
    ]

    # ========== RÉPONSE FINALE ==========
    return Response({
        'overview': {
            'totalEvents': all_events.count(),
            'activeEvents': active_events.count(),
            'upcomingEvents': upcoming_events.count(),
            'pastEvents': past_events.count(),
            'totalParticipants': total_participants,
            'newParticipantsThisMonth': new_this_month,
            'totalBungalows': total_bungalows,
            'occupiedBungalows': occupied_bungalows,
            'totalBedCapacity': total_bed_capacity
        },
        'events': {
            'byType': {
                'stages': event_type_counts.get('stage', 0),
                'residences': event_type_counts.get('resident', 0),
                'autres': event_type_counts.get('autres', 0)
            },
            'active': active_events_list,
            'upcoming': upcoming_events_list
        },
        'participants': {
            'byStatus': {
                'students': status_counts.get('student', 0),
                'instructors': status_counts.get('instructor', 0),
                'professionals': status_counts.get('professional', 0),
                'staff': status_counts.get('staff', 0)
            },
            'byGender': {
                'men': gender_counts.get('M', 0),
                'women': gender_counts.get('F', 0)
            },
            'averageAge': average_age,
            'topNationalities': top_nationalities,
            'topLanguages': top_languages
        },
        'registrations': {
            'activeTotal': total_active_registrations,
            'assigned': assigned_registrations,
            'unassigned': unassigned_registrations,
            'byRole': {
                'participants': role_counts.get('participant', 0),
                'instructors': role_counts.get('instructor', 0),
                'musicians': role_counts.get('musician', 0),
                'staff': role_counts.get('staff', 0)
            }
        },
        'villages': village_stats,
        'alerts': alerts,
        'recentActivities': activities_list,
        'trends': {
            'participantsTrend': round(participant_trend, 1),
            'eventsTrend': round(event_trend, 1),
            'newParticipantsThisMonth': new_this_month,
            'newParticipantsLastMonth': new_last_month,
            'eventsThisMonth': events_this_month,
            'eventsLastMonth': events_last_month
        },
        'lastUpdated': timezone.now().isoformat()
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def auto_assign_stage_participants(request, stage_id):
    """
    Assigne automatiquement tous les participants non assignés d'un stage aux bungalows.

    Logique d'assignation:
    1. Encadrants → chambre individuelle + lit double si possible
    2. Staff → chambre individuelle si possible
    3. Musiciens/Participants → optimisation du remplissage (grouper par genre)

    POST /api/stages/<stage_id>/auto-assign/
    """
    try:
        stage = Stage.objects.get(pk=stage_id)
    except Stage.DoesNotExist:
        return Response({
            'success': False,
            'error': f"Stage avec ID {stage_id} non trouvé"
        }, status=status.HTTP_404_NOT_FOUND)

    try:
        # Lancer l'assignation automatique
        results = assign_participants_automatically_for_stage(stage)

        # Compter les résultats
        success_count = len(results['success'])
        failure_count = len(results['failure'])

        # Log individuel pour chaque assignation réussie
        for assignment in results['success']:
            log_auto_assignment_individual(
                user=request.user,
                participant_name=assignment['participant'],
                stage_name=assignment['stage'],
                bungalow_name=assignment['bungalow'],
                bed_id=assignment['bed'],
                village_name=assignment.get('village')
            )

        # Log résumé de l'assignation automatique
        if success_count > 0 or failure_count > 0:
            log_auto_assignment_summary(request.user, stage.name, success_count, failure_count)

        return Response({
            'success': True,
            'stage': {
                'id': stage.id,
                'name': stage.name
            },
            'summary': {
                'total_assigned': success_count,
                'total_failed': failure_count,
                'success_rate': round((success_count / (success_count + failure_count) * 100), 1) if (success_count + failure_count) > 0 else 0
            },
            'assignments': results['success'],
            'failures': results['failure'],
            'message': f"Assignation automatique terminée: {success_count} participant(s) assigné(s), {failure_count} échec(s)"
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            'success': False,
            'error': f"Erreur lors de l'assignation automatique: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== NETWORK INFO ====================

@api_view(['GET'])
@permission_classes([])  # Pas d'authentification requise
def network_info(request):
    """Retourne l'IP réseau locale du serveur."""
    import socket

    try:
        # Même méthode que dans manage.py
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()

        return Response({
            'local_ip': local_ip,
            'success': True
        })
    except Exception as e:
        return Response({
            'local_ip': '127.0.0.1',
            'success': False,
            'error': str(e)
        })
